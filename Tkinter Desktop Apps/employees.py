from tkinter import*
from tkinter import ttk
from tkinter.messagebox import*
from tkinter.messagebox import*
import openpyxl
from openpyxl import Workbook 
import pyttsx3
#--------- speech word-------------
word=pyttsx3.init()
word.say(" Devloper AbdalRahman  welcome you in his Application")
word.setProperty("rate",150) #للتحكم فى سرعة ومعدل الصوت
word.runAndWait()
#----------------- Adding data to Excel--------------------

wb=Workbook()     # لأنشاء ملف اكسلworkbookلأستدعاء 
ws=wb.active      #للتنشيط

ws.title = " < Customer-Fatora > "    # عنوان ملف اكسل
ws["A1"]=' Name'                  # الاعمدة بعنوينهابالترتيب
ws["B1"]='Job '
ws["C1"]='gender'
ws['D1']='Age '
ws["E1"]='Email'
ws["F1"]='mobile'
ws["G1"]='Address'


wb.save('emloyee.xlsx')

#---------------- objects----------------

root=Tk()
root.title("employees")

root.iconbitmap("C:\\Users\\EL-Hussein Store\\Desktop\\images\\c2.ico")
root.geometry("1310x600+15+20")

root.config(background="whitesmoke")
root.resizable(False,False)

#----------- متغيرات---------------
v1=StringVar()
v2=StringVar()
v3=StringVar()
v4=IntVar()
v5=StringVar()
v6=IntVar()
v7=StringVar()


#------------------- frame1--------------------
fr1=Frame(root,width=300,height=400,bg="#A6F7AA",bd=2,relief=GROOVE).place(x=1,y=1)
title=Label(fr1,text="*-> Employees-Company <-*",bg="#A6F7AA",fg="#E74C3C",
            cursor="hand2",font=("Calibri",18,"bold")).place(x=10,y=1)

name=Label(fr1,text="->>الأسم:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15)).place(x=10,y=50)

en1=Entry(fr1,width=18,fg="black",font=("Calibri",15),textvariable=v1,justify="center",cursor="hand2").place(x=100,y=50)


job=Label(fr1,text="->>الوظيفة:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15)).place(x=10,y=100)

en2=Entry(fr1,width=18,fg="black",font=("Calibri",15),textvariable=v2,justify="center",cursor="hand2").place(x=100,y=100)


gender=Label(fr1,text="->>أختر جنسك:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",12)).place(x=10,y=150)

en3=ttk.Combobox(fr1,values=("ذكر","أنثى"),textvariable=v3,font=("Calibri",12),width=7,state='readonly',cursor="hand2").place(x=120,y=150)

Age=Label(fr1,text="->>العمر:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15)).place(x=10,y=200)

en4=Spinbox(fr1,text="->>أدخل عمرك",from_=0,to=100,textvariable=v4,fg="black",font=("Tajawal",12),width=7,cursor="hand2").place(x=100,y=200)


Email=Label(fr1,text="->>ألأيميل ألألكترونى:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15),width=10).place(x=10,y=250)

en5=Entry(fr1,width=14,fg="black",font=("Calibri",15),textvariable=v5,justify="center",cursor="hand2").place(x=130,y=250)


mobile=Label(fr1,text="->>رقم الهاتف:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15)).place(x=10,y=300)

en6=Entry(fr1,width=16,fg="black",font=("Calibri",15),justify="center",textvariable=v6,cursor="hand2").place(x=120,y=300)

address=Label(fr1,text="->> العنوان:",bg="#A6F7AA",fg="#0000FF",
            cursor="hand2",font=("Calibri",15)).place(x=10,y=350)
en7=ttk.Combobox(fr1,width=15,font=("Calibri",8),cursor="hand2",textvariable=v7,values=("cairo","Aswan","Giza","Alxor","Alminia","alqulupia","Helwan","Alfayoum","Almohandessen","Alex"))
en7.place(x=100,y=360)

image7=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\exit3.png")
img7=image7.subsample(7,7)

image8=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\enter.png")
img8=image8.subsample(7,7)
#----------- متغيرات---------------

def save():
      q=askokcancel("الحفظ"," هل تريد ان يتم حفظ الملف")
      if q==0:
          return
      else:
          nam=en1.get()
          jo=en2.get()
          gend=en3.get()
          Ag=en4.get()
          Ema=en5.get()
          mo=en6.get()
          add=en7.get()

          excel = openpyxl.load_workbook("employee.xlsx")
          file=excel.active
          file.cell(column=1,row=file.max_row+1,value=nam)
          file.cell(column=2,row=file.max_row,value=jo)
          file.cell(column=3,row=file.max_row,value=gend)
          file.cell(column=4,row=file.max_row,value=Ag)
          file.cell(column=5,row=file.max_row,value=Ema)
          file.cell(column=5,row=file.max_row,value=mo)
          file.cell(column=5,row=file.max_row,value=add)
          excel.save("employee.xlsx")

#------------- الجزء الخاص بزر الأخفاء--------------
def on_enter11(event=None):
    hide_btn["background"]="tomato1"
def on_leave11(event=None):
    hide_btn["background"]="whitesmoke"    

def hide(event=None):
    root.geometry("300x600+15+20")
    
hide_btn=Button(fr1,bg="#A6F7AA",cursor="hand2",image=img8,command=hide,relief=GROOVE,bd=1)
hide_btn.place(x=210,y=175)

hide_btn.bind('<Enter>',on_enter11)
hide_btn.bind('<Leave>',on_leave11)

#------------- الجزء الخاص بزر الظهور--------------
def on_enter12(event=None):
    show_btn["background"]="#0000EE"
def on_leave12(event=None):
    show_btn["background"]="whitesmoke"    

def show(event=None):
    root.geometry("1310x600+15+20")
    
show_btn=Button(fr1,bg="#A6F7AA",cursor="hand2",image=img7,command=show,relief=GROOVE,bd=1)
show_btn.place(x=250,y=175)

show_btn.bind('<Enter>',on_enter12)
show_btn.bind('<Leave>',on_leave12)
#------------- insert into treeview--------------
def add():
    for item in trv.get_children():
      item.insert('1',str(en1) )
      item.insert('1',str(en2))
 
#----------- و حقول الأدخالtreeview  افراغ ------------
def clear():
    for item in trv.get_children():
        trv.delete(item)
    en1.delete('0',END)
    en2.delete('0',END)
    en3.delete('0',END)
    en4.delete('0',END)
    en5.delete('0',END)
    en6.delete('0',END)
    en7.delete('0',END)
   
#---------------  delete function----------------------
def delete():
        for item in trv .get_children():
           trv.delete(item)

#------------------- All important images---------------------------------
image2=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\employee.png")
img2=image2.subsample(1,2)

image3=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\plus.png")
img3=image3.subsample(6,6)

image4=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\update.png")
img4=image4.subsample(5,5)

image5=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\syper1.png")
img5=image5.subsample(5,5)

image6=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\clear.png")
img6=image6.subsample(4,4)

#--------------- All honvor for buttons--------------------
def on_enter(event=None):
    add_btn["background"]="#778899"
def on_leave(event=None):
    add_btn["background"]="whitesmoke"    

def on_enter1(event=None):
    update_btn["background"]="#778899"
def on_leave1(event=None):
    update_btn["background"]="whitesmoke"  

def on_enter2(event=None):
    save_btn["background"]="#778899"
def on_leave2(event=None):
    save_btn["background"]="whitesmoke" 

def on_enter3(event=None):
    clear_btn["background"]="#778899"
def on_leave3(event=None):
    clear_btn["background"]="whitesmoke" 

#-------------------- Buttons + frame -------------------------
fr2=Frame(root,width=300,height=200,bg="#A6F7AA",bd=2,relief=GROOVE).place(x=1,y=400)

add_btn=Button(fr2,text="add-Details",font=("Tajawal",13),cursor="hand2",width=100
           ,bg="whitesmoke",fg="#FF6347",activebackground="whitesmoke",
           activeforeground="#FF3030",bd=1,relief=SOLID,image=img3,compound=LEFT,command=add)

add_btn.place(x=5,y=405)
add_btn.bind('<Enter>',on_enter)
add_btn.bind('<Leave>',on_leave)

update_btn=Button(fr2,text="update",font=("Tajawal",13),cursor="hand2",width=100
           ,bg="whitesmoke",fg="#FF6347",activebackground="whitesmoke",
           activeforeground="#FF3030",bd=1,relief=SOLID,compound=LEFT,image=img4)

update_btn.place(x=150,y=405)
update_btn.bind('<Enter>',on_enter1)
update_btn.bind('<Leave>',on_leave1)

save_btn=Button(fr2,text="save",font=("Tajawal",13),cursor="hand2",width=100
           ,bg="white",fg="#FF6347",activebackground="whitesmoke",
           activeforeground="#FF3030",bd=1,relief=SOLID,compound=LEFT,image=img5,command=save)

save_btn.place(x=3,y=460)
save_btn.bind('<Enter>',on_enter2)
save_btn.bind('<Leave>',on_leave2)

clear_btn=Button(fr2,font=("Tajawal",13),cursor="hand2",width=100
           ,bg="white",fg="#FF6347",activebackground="whitesmoke",
           activeforeground="#FF3030",bd=1,relief=SOLID,image=img6,command=clear)

clear_btn.place(x=150,y=480)
clear_btn.bind('<Enter>',on_enter3)
clear_btn.bind('<Leave>',on_leave3)

logo=Label(fr2,image=img2,cursor="hand2",bg="#A6F7AA")
logo.place(x=50,y=525)

#--------------main-title-------------------
title=Label(root,text="*->>  نظام ادارة الموظفين   <<-*",font=("Tajwal",13,"bold")
            ,fg="whitesmoke",bg="#1A1A1A",bd=1,width=100).place(x=300,y=5)

#-------------- treeview (Table Frame )-----------------------
trv=ttk.Treeview(root,selectmode="browse")
trv.place(x=310,y=30,width=990,height=580)

#----------------- create coulumns-----------------

trv["columns"]=("1","2","3","4","5","6","7")
trv.column('#0',width=25,anchor='c')
trv.column('1',width=25,anchor='c')
trv.column('2',width=25,anchor='c')
trv.column('3',width=25,anchor='c')
trv.column('4',width=25,anchor='c')
trv.column('5',width=25,anchor='c')
trv.column('6',width=25,anchor='c')
trv.column('7',width=25,anchor='c')

#---------------- create headings---------------------

trv.heading("#0",text="ID",anchor="c")
trv.heading("1",text="Name",anchor="c")
trv.heading("2",text="Age",anchor="c")
trv.heading("3",text="Job",anchor="c")
trv.heading("4",text="Email",anchor="c")
trv.heading("5",text="Gender",anchor="c")
trv.heading("6",text="mobile",anchor="c")
trv.heading("7",text="Address",anchor="c")


root.bind('<Control-v>',hide)
root.mainloop()