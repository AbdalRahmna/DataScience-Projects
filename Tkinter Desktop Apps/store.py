from tkinter import*
import datetime  #pkg to mange in time and date
from tkinter import ttk
import openpyxl  # to dealing with excel files
from openpyxl import Workbook  #  to create excel fiel
from tkinter.messagebox import*
import pyttsx3
#--------- speech word-------------
word=pyttsx3.init()
word.say(" Devloper AbdalRahman  welcome you in his Application")
word.setProperty("rate",150) #للتحكم فى سرعة ومعدل الصوت
word.runAndWait()
#-----objects------
root=Tk()
root.title("[متجر أدوات بنائية]")
root.iconbitmap("C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\الايقونة-البنا.ico")
root.geometry("1024x700+5+3")
root.resizable(False,False)
root.config(background="whitesmoke")

now=datetime.datetime.now()     #جبلى التاريخ والوقت من الحاسوب
date=now.strftime("%Y-%m-%d")

#----------------- Adding data to Excel--------------------

wb=Workbook()     # لأنشاء ملف اكسلworkbookلأستدعاء 
ws=wb.active      #للتنشيط

ws.title = " < Customer-Fatora > "    # عنوان ملف اكسل
ws["A1"]='الأسم بالكامل '                  # الاعمدة بعنوينهابالترتيب
ws["B1"]=' التليفون'
ws["C1"]='العنوان'
ws['D1']='الحساب الكلى'
ws["E1"]='التاريخ'
wb.save('bulding_store.xlsx')

#------------ prices-----------
menu={
    0:['برويطة',30],
    1:['حفار',5],
    2:['حفارة',100],
    3:['طوب',2],
    4:['بندوزر',50],
    5:['دهان',120],
    6:['مترين',30],
    7:['مسطرين',20],
    8:['مفكات',5],
    9:['منشار',12],
    10:['مطرقة',15],
    11:['كلاون',150],
    12:['حديد',1050],
    13:['أسمنت',900],
    14:['شوكة قش',20]
   }

def close():
    p=askokcancel("أغلاق","هل تريد أغلاق البرنامج؟")
    if p==0:
        return
    else:
        root.quit()

def hide(event=None):
    root.geometry("1024x700+50+20")

def bill():
    #--------- متغيرات عامة-----------
    global name
    global phon
    global address
    global  total
    global Date

    root.geometry("1350x900+5+3")
    pg.place(x=1020,y=500)
    fr3=Frame(root,bg="#5F7161",width=400,height=500,bd=2,relief=GROOVE)
    fr3.place(x=1020,y=1)
    
    name=Label(fr3,text=":اسم المشترى-",bg="#5F7161",fg="red",font=("Tajwal",13))
    name.place(x=190,y=10)
    
    En1=Entry(fr3,width=20,bg="whitesmoke",cursor="hand2",font=("Tajawal",12),justify="center")
    En1.place(x=4,y=10)

    phon=Label(fr3,text=": رقم الهاتف-",bg="#5F7161",fg="red",font=("Tajwal",13))
    phon.place(x=200,y=80)
    
    En2=Entry(fr3,width=20,bg="whitesmoke",cursor="hand2",font=("Tajawal",12),justify="center")
    En2.place(x=4,y=80)
    
    address=Label(fr3,text=": عنوان المشترى-",bg="#5F7161",fg="red",font=("Tajwal",13))
    address.place(x=190,y=150)
    
    
    En3=Entry(fr3,width=20,bg="whitesmoke",cursor="hand2",font=("Tajawal",12),justify="center")
    En3.place(x=4,y=150)
    
    total=Label(fr3,text=": الحساب الأجمالى-",bg="#5F7161",fg="red",font=("Tajwal",13))
    total.place(x=190,y=220)
    
    En4=Entry(fr3,width=20,bg="whitesmoke",cursor="hand2",font=("Tajawal",12),justify="center")
    En4.place(x=4,y=220)

    Date=Label(fr3,text=": تاريخ الشراء-",bg="#5F7161",fg="red",font=("Tajwal",13))
    Date.place(x=190,y=280)
    
    En5=Entry(fr3,width=20,bg="whitesmoke",cursor="hand2",font=("Tajawal",12),justify="center")
    En5.place(x=4,y=280)
    
#-------------------------------------------------
    def on_enter11(e):
         add_btn["background"]="green"
    
    def on_leave11(e):
        add_btn["background"]="gray"
    
    #------------- دالة أفرغ الحقول--------------
    def empty():
        En1.delete('0',END)
        En2.delete('0',END)
        En3.delete('0',END)
        En4.delete('0',END)
        En5.delete('0',END)
    #---------- باقى الجزء الخاص بالالاكسل-----------------
    def save():
      q=askokcancel("الحفظ","هل تريد أن تحفظ الفاتورة؟")
      if q==0:
          return
      else:
          nam=En1.get()
          pho=En2.get()
          add=En3.get()
          tot=En4.get()
          datbuy=En5.get()

          excel = openpyxl.load_workbook("bulding_store.xlsx")
          file=excel.active
          file.cell(column=1,row=file.max_row+1,value=nam)
          file.cell(column=2,row=file.max_row,value=pho)
          file.cell(column=3,row=file.max_row,value=add)
          file.cell(column=4,row=file.max_row,value=tot)
          file.cell(column=5,row=file.max_row,value=datbuy)
          excel.save("bulding_store.xlsx")
    #---------- دالة من  نحن------------
    def about():
        showinfo("المطور","تم برمجة البرنامج بواسطة\nDevloper:AbdalRahman Gameel Hebishy")
    
    #----------- دالة حذف فاتورة---------------
    def delete():
        for item in trv .get_children():
           trv.delete(item)
        En1.delete('0',END)
        En2.delete('0',END)
        En2.delete('0',END)
        En2.delete('0',END)
        En2.delete('0',END)
    #---------------------honvor الأزرار مع ---------------
    add_btn=Button(fr3,text=" حفظ الفاتورة",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=save)
    add_btn.place(x=30,y=350)
    
    add_btn.bind('<Enter>',on_enter11)
    add_btn.bind('<Leave>',on_leave11)
    
    def on_enter12(e):
        exit_btn["background"]="green"
    
    def on_leave12(e):
        exit_btn["background"]="gray"
    
    
    exit_btn=Button(fr3,text="أخفاء",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=hide)
    
    exit_btn.place(x=180,y=350)
     
    exit_btn.bind('<Enter>',on_enter12) 
    exit_btn.bind('<Leave>',on_leave12) 

    def on_enter13(e):
        empty_btn["background"]="green"
    
    def on_leave13(e):
        empty_btn["background"]="gray"
    
    empty_btn=Button(fr3,text=":أفراغ الحقول",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=empty)
    
    empty_btn.place(x=30,y=400)
    
    empty_btn.bind('<Enter>',on_enter13) 
    empty_btn.bind('<Leave>',on_leave13)
    

    def on_enter14(e):
        search_btn["background"]="green"
    
    def on_leave14(e):
        search_btn["background"]="gray"

    search_btn=Button(fr3,text=":   المطور",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=about)
    search_btn.place(x=180,y=400)

    search_btn.bind('<Enter>',on_enter14) 
    search_btn.bind('<Leave>',on_leave14)
    
    def on_enter15(e):
        delete_btn["background"]="green"
    
    def on_leave15(e):
        delete_btn["background"]="gray"
    
    delete_btn=Button(fr3,text=":  حذف الفاتورة ",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=delete)
    delete_btn.place(x=30,y=450)

    delete_btn.bind('<Enter>',on_enter15) 
    delete_btn.bind('<Leave>',on_leave15)
    
    def on_enter16(e):
        close2_btn["background"]="green"
    
    def on_leave16(e):
        close2_btn["background"]="gray"
    
    close2_btn=Button(fr3,text=":   المغادرة ",width=10,
                   font=("Tajawal",13),cursor="hand2",
                   bg="gray",fg="whitesmoke",activebackground="yellow",
                   activeforeground="green",bd=2,relief=SOLID,command=fr3.quit)
    close2_btn.place(x=180,y=450)

    close2_btn.bind('<Enter>',on_enter16) 
    close2_btn.bind('<Leave>',on_leave16)
    
    #--------------- الشجرة وما فى داخلها ------------
    total=0

    for item in trv.get_children():    # لبعث متغير الى الشجرة والبحث بداخلها
           trv.delete(item)
    
    for i in range(len(list)) :
        if(int(list[i].get())>0) :
            price=int(list[i].get())*menu[i][1]
            total=total+price
            myst=(str(menu[i][1]), str(list[i].get()) , str(price))
            trv.insert("",'end',iid=i,text=menu[i][0],values=myst)

    finall=total
    En4.insert('1',str(finall) + "$")
    En5.insert('1',str(date))
 
    
#----------- و حقول الأدخالtreeview  افراغ ------------
def clear():
    for item in trv .get_children():
        trv.delete(item)
    
#--------------- All  images used ----------------------
image1=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\برويطة.png")
img1=image1.subsample(2,2)

image2=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\حفار.png")
img2=image2.subsample(2,2)

image3=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\images\\عناصر المتجر\\حفارة.png")
img3=image3.subsample(2,2)

image4=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\طوب.png")
img4=image4.subsample(2,2)

image5=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\فرشاة دهان.png")
img5=image5.subsample(2,2)

image6=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\مترين.png")
img6=image6.subsample(2,2)

image7=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\مسطرين.png")
img7=image7.subsample(2,2)

image8=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\مفك.png")
img8=image8.subsample(2,2)

image9=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\منشار.png")
img9=image9.subsample(2,2)

image10=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\Bundozer.png")
img10=image10.subsample(2,2)

image11=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\مطارق.png")
img11=image11.subsample(2,2)

image12=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\كوالين.png")
img12=image12.subsample(2,2)

image13=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\حديد.png")
img13=image13.subsample(2,2)

image14=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\أسمنت.png")
img14=image14.subsample(2,2)

image15=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\شوكة-قش.png")
img15=image15.subsample(2,2)

image16=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\الحساب2.png")
img16=image16.subsample(6,6)

image17=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\fatora.png")
img17=image17.subsample(6,6)

image18=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\استاجار.png")
img18=image18.subsample(6,6)

image19=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\exit2.png")
img19=image19.subsample(6,6)


image20=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\خلفية-استئجار-مواد-بنائية.png")


#-------------- frame1-----------------------------


fr1=Frame(root,width=630,height=675,bg="silver")
fr1.place(x=1,y=1)


#---------All Labels-----------------

title=Label(fr1,text="[مشروع بيع أدوات البناء:(المعدات والأدوات المتوفرة)]",
            fg="#0000EE"
            ,cursor="hand2",bd=1,font=("Tajawal",16,"bold"),bg="#ADF7A6")
title.place(x=150,y=10)


menu1=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img1,text="برويطة",compound=TOP).place(x=30,y=80)


menu2=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img2,text="حفار",compound=TOP).place(x=150,y=80)


menu3=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img3,text="حفارة",compound=TOP).place(x=270,y=80)


menu4=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img4,text="طوب بأنواعة",compound=TOP).place(x=390,y=80)


menu5=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img5,text=" دهان وأدواتة",compound=TOP).place(x=30,y=250)


menu6=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img6,text=" مترين",compound=TOP).place(x=160,y=250)


menu7=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img7,text=" مسطرين",compound=TOP).place(x=280,y=250)


menu8=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img8,text=" مفكات بأنواعها",compound=TOP).place(x=400,y=250)


menu9=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img9,text=" مناشير",compound=TOP).place(x=520,y=250)


menu10=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img10,text=" بندوذر",compound=TOP).place(x=510,y=80)


menu11=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img11,text=" مطارق بأنواعها",compound=TOP).place(x=30,y=420)


menu12=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img12,text=" كولاين",compound=TOP).place(x=150,y=420)


menu13=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img13,text=" حديد",compound=TOP).place(x=270,y=420)


menu14=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img14,text=" أسمنت",compound=TOP).place(x=390,y=420)


menu15=Label(fr1,width=88,bg="#F1F7A6",relief=SOLID,
            cursor="hand2",image=img15,text=" شوكة قش",compound=TOP).place(x=510,y=420)



#--------------------- All variables-------------------------
list=[]
font1=("Times",12,"normal")

v1=IntVar()
v2=IntVar()
v3=IntVar()
v4=IntVar()
v5=IntVar()
v6=IntVar()
v7=IntVar()
v8=IntVar()
v9=IntVar()
v10=IntVar()
v11=IntVar()
v12=IntVar()
v13=IntVar()
v14=IntVar()
v15=IntVar()


#------------ Enteries(spinbox deal with only numbers) ---------------------------

en1=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000
            ,font=font1,textvariable=v1,fg="red",cursor="hand2")
en1.place(x=30,y=220)
list.append(en1)


en2=Spinbox(fr1,width=10,bg="whitesmoke",font=font1
            ,textvariable=v2,from_=0,to=1000,fg="red",cursor="hand2")
en2.place(x=150,y=220)
list.append(en2)


en3=Spinbox(fr1,width=10,bg="whitesmoke",font=font1,textvariable=v3
            ,from_=0,to=1000,fg="red",cursor="hand2")
en3.place(x=270,y=220)
list.append(en3)


en4=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,
            font=font1,textvariable=v4,to=1000,fg="red",cursor="hand2")
en4.place(x=390,y=220)
list.append(en4)


en5=Spinbox(fr1,width=10,bg="whitesmoke",
            font=font1,textvariable=v5,from_=0,to=1000,fg="red",cursor="hand2")
en5.place(x=510,y=220)
list.append(en5)


en6=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,
            font=font1,textvariable=v6,fg="red",cursor="hand2")
en6.place(x=40,y=390)
list.append(en6)


en7=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,fg="red",
            font=font1,textvariable=v7,cursor="hand2")
en7.place(x=160,y=390)
list.append(en7)


en8=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,
            font=font1,textvariable=v8,fg="red",cursor="hand2")
en8.place(x=280,y=390)
list.append(en8)


en9=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,fg="red",
            font=font1,textvariable=v9,cursor="hand2")
en9.place(x=400,y=390)
list.append(en9)


en10=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,fg="red",
             font=font1,textvariable=v10,cursor="hand2")
en10.place(x=520,y=390)
list.append(en10)

en11=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,
             font=font1,textvariable=v11,
             fg="red",cursor="hand2")
en11.place(x=30,y=550)
list.append(en11)

en12=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,font=font1,
             textvariable=v12,to=1000,fg="red",cursor="hand2")
en12.place(x=150,y=550)
list.append(en12)

en13=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,
             font=font1,textvariable=v13,fg="red",cursor="hand2")
en13.place(x=270,y=550)
list.append(en13)

en14=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,
             font=font1,textvariable=v14,fg="red",cursor="hand2")
en14.place(x=390,y=570)
list.append(en14)

en15=Spinbox(fr1,width=10,bg="whitesmoke",from_=0,to=1000,fg="red",
             font=font1,textvariable=v15,
             cursor="hand2")
en15.place(x=510,y=560)
list.append(en15)

#--------------honvor الدوال المسئولة عن ----------------
def on_enter(e):
    btn1["background"]="red" #color here shouldnt be the same color for background btn 

def on_leave(e):
    btn1["background"]='silver'    

def on_enter1(e):
    btn2["background"]="red"

def on_leave1(e):
    btn2["background"]='silver'    

def on_enter2(e):
    btn2["background"]="red"

def on_leave2(e):
    btn2["background"]='silver' 

def on_enter3(e):
    btn3["background"]="red"

def on_leave3(e):
    btn3["background"]="silver"

def on_enter4(e):
    btn4["background"]="red"

def on_leave4(e) :
    btn4["background"]="silver"
#----------------- دالة  الاستئجار------------------------
def astager():
    showinfo("الأستئجار","حاليا لاتوجد ألأيات يمكن تاجيرها")

#----------------- Buttons---------------------

btn1=Button(fr1,text="شراء المواد",fg="whitesmoke",
            font=("Tajawal",12),bg="silver",bd=1,
            relief=SOLID,cursor="hand2",activebackground="yellow",command=bill,activeforeground="pink",image=img16,compound=TOP)
btn1.place(x=30,y=600)

btn1.bind('<Enter>',on_enter)
btn1.bind('<Leave>',on_leave)

btn2=Button(fr1,text=" فاتورة جديدة",fg="whitesmoke",
            font=("Tajawal",12),bg="silver",bd=1,
            relief=SOLID,cursor="hand2",
            activebackground="yellow",activeforeground="pink",image=img17,compound=TOP,command=clear)
btn2.place(x=160,y=600)

btn2.bind('<Enter>',on_enter1)
btn2.bind("<Leave>",on_leave1)

btn3=Button(fr1,text="  أستجار معدات البناء",fg="whitesmoke",
            font=("Tajawal",12),bg="silver",bd=1,
            relief=SOLID,cursor="hand2",
            activebackground="yellow",activeforeground="pink",
            image=img18,compound=TOP,command=astager)
btn3.place(x=290,y=605)

btn3.bind('<Enter>',on_enter3)
btn3.bind("<Leave>",on_leave3)

btn4=Button(fr1,text="أغلاق البرنامج",fg="whitesmoke",
            font=("Tajawal",12),bg="silver"
            ,cursor="hand2",
            activebackground="yellow",activeforeground="pink",
            image=img19,compound=LEFT,command=close)
btn4.place(x=420,y=610)

btn4.bind("<Enter>",on_enter4)
btn4.bind("<Leave>",on_leave4)

#-----------------  frame2 -------------------

fr2=Frame(root,width=385,height=400,bd=2,relief=GROOVE,bg="whitesmoke")
fr2.place(x=635,y=1)

#---------------- Treeview -------------------
trv=ttk.Treeview(fr2,selectmode="browse")
trv.place(width=385,height=675,x=1,y=1)

trv['columns']=('1','2','3')

trv.column("#0",width=80,anchor='c')
trv.column("1",width=50,anchor='c')
trv.column("2",width=50,anchor='c')
trv.column("3",width=60,anchor='c')

trv.heading("#0",text="المواد",anchor='c')
trv.heading("1",text="السعر",anchor='c')
trv.heading("2",text="العدد",anchor='c')
trv.heading("3",text="الحساب الكلى",anchor='c')


root.bind('<Control-v>',hide)

logo_image=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\market logo.png")

pg=Label(root,image=logo_image,cursor="hand2",
         text="*-*-*-*-*-*-*-*-*-*\n*-*-*-*-*-*-*-*-*\nBUlding-Store\n*-*-*-*-*-*-*-*-*\n*-*-*-*-*-*-*-*-*-*",font=("Tajawal",14)
         ,width=400,bg="#BAF7A6",compound=RIGHT)

logo=PhotoImage(file="C:\\Users\\EL-Hussein Store\\Desktop\\images\\عناصر المتجر\\الايقونة-البنا.png")

pht=Label(root,image=logo,cursor="hand2",width=370,height=280,bg="whitesmoke",bd=1,relief=SOLID)
pht.place(x=640,y=400)

root.mainloop()